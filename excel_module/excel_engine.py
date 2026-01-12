from __future__ import annotations

import json
import logging
import tempfile
import shutil
from pathlib import Path
from typing import Optional, Dict, Any, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import pythoncom
import win32com.client as win32

logger = logging.getLogger("prova.excel_engine")
logger.addHandler(logging.NullHandler())

# =====================================================
# Constants
# =====================================================
MASTER_SHEET = "Master"
DASHBOARD_SHEET = "Dashboard"
CONFIG_SHEET = "Config"

XL_CELL_TYPE_LAST_CELL = 11
XL_ROW_FIELD = 1

# Pivot aggregation
XL_SUM = -4157
XL_COUNT = -4112
XL_AVERAGE = -4106
XL_MAX = -4136
XL_MIN = -4139

_AGG_MAP = {
    "sum": XL_SUM,
    "count": XL_COUNT,
    "avg": XL_AVERAGE,
    "average": XL_AVERAGE,
    "max": XL_MAX,
    "min": XL_MIN,
}

# Chart types
XL_CHART_TYPES = {
    "line": 4,
    "bar": 57,
    "column": 51,
    "pie": 5,
    "area": 1,
    "xlline": 4,
    "xlbar": 57,
    "xlbarclustered": 57,
    "xlcolumn": 51,
    "xlcolumnclustered": 51,
}

CHART_WIDTH = 420
CHART_HEIGHT = 260

# =====================================================
# Themes
# =====================================================
THEMES = {
    "softblue": {
        "primary": 0xBD814F,
        "secondary": 0x9CC3E5,
        "font": "Calibri",
        "font_size": 11,
    },
    "dark": {
        "primary": 0x00B0F0,
        "secondary": 0x7030A0,
        "font": "Calibri",
        "font_size": 11,
    },
}

DEFAULT_CONFIG = {
    "theme": "softblue",
    "legend_position": "right",
    "charts": {
        "show_legend": True,
        "gridlines": False,
        "line_width": 2,
        "font_size": 11,
        "show_data_labels": False,
    },
    "kpis": {
        "font_size": 20,
        "positive_color": "#2ECC71",
        "negative_color": "#E74C3C",
        "neutral_color": "#34495E",
    },
    "slicer_style": "SlicerStyleLight1",
}

# =====================================================
# Helpers
# =====================================================
def hex_to_rgb_int(hex_color: str) -> int:
    hex_color = hex_color.lstrip("#")
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return b << 16 | g << 8 | r


def read_config(template: Path) -> dict:
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))  # deep copy
    try:
        wb = load_workbook(template, read_only=True, data_only=True)
        if CONFIG_SHEET in wb.sheetnames:
            raw = wb[CONFIG_SHEET].cell(2, 1).value
            if isinstance(raw, str):
                cfg.update(json.loads(raw))
    except Exception:
        pass
    return cfg


# =====================================================
# Excel lifecycle
# =====================================================
def open_excel(visible: bool):
    pythoncom.CoInitialize()
    xl = win32.DispatchEx("Excel.Application")
    xl.Visible = visible
    xl.DisplayAlerts = False
    xl.ScreenUpdating = False
    return xl


def close_excel(xl):
    try:
        xl.Quit()
    except Exception:
        pass
    pythoncom.CoUninitialize()


# =====================================================
# Pivot helpers
# =====================================================
def create_pivot_cache(wb, src: str):
    return wb.PivotCaches().Create(SourceType=1, SourceData=src)


def add_pivot(wb, cache, spec: Dict[str, Any]):
    sheet = wb.Worksheets.Add()
    sheet.Name = spec["sheet"]

    pvt = cache.CreatePivotTable(sheet.Cells(1, 1), spec["name"])

    for r in spec.get("rows", []):
        pvt.PivotFields(r).Orientation = XL_ROW_FIELD

    for v, agg in spec.get("values", []):
        func = _AGG_MAP[str(agg).lower()]
        pvt.AddDataField(
            pvt.PivotFields(v),
            f"{agg.capitalize()} of {v}",
            func,
        )

    return pvt


# =====================================================
# KPI styling
# =====================================================
def style_kpi_cell(cell, value, cfg):
    kpi_cfg = cfg.get("kpis", {})
    cell.Font.Bold = True
    cell.Font.Size = kpi_cfg.get("font_size", 20)

    if value > 0:
        color = kpi_cfg.get("positive_color", "#2ECC71")
    elif value < 0:
        color = kpi_cfg.get("negative_color", "#E74C3C")
    else:
        color = kpi_cfg.get("neutral_color", "#34495E")

    cell.Font.Color = hex_to_rgb_int(color)


# =====================================================
# Chart styling
# =====================================================
def style_chart(chart, cfg: dict, title: Optional[str]):
    charts_cfg = cfg.get("charts", {})
    theme = THEMES.get(cfg.get("theme", "softblue"), THEMES["softblue"])

    if title:
        chart.HasTitle = True
        chart.ChartTitle.Text = title

    chart.HasLegend = charts_cfg.get("show_legend", True)

    if chart.HasLegend:
        try:
            chart.Legend.Position = {
                "right": 2,
                "left": 1,
                "top": 0,
                "bottom": 3,
            }.get(cfg.get("legend_position", "right"), 2)
        except Exception:
            pass

    try:
        chart.Axes(1).HasMajorGridlines = charts_cfg.get("gridlines", False)
        chart.Axes(2).HasMajorGridlines = charts_cfg.get("gridlines", False)
    except Exception:
        pass

    try:
        for i in range(1, chart.SeriesCollection().Count + 1):
            s = chart.SeriesCollection(i)
            s.Format.Line.Visible = True
            s.Format.Line.Weight = charts_cfg.get("line_width", 2)
            s.Format.Line.ForeColor.RGB = theme["primary"]
            s.HasDataLabels = charts_cfg.get("show_data_labels", False)
    except Exception:
        pass

    try:
        tr = chart.ChartArea.Format.TextFrame2.TextRange
        tr.Font.Name = theme["font"]
        tr.Font.Size = charts_cfg.get("font_size", theme["font_size"])
    except Exception:
        pass

    try:
        chart.ShowAllFieldButtons = False
    except Exception:
        pass


# =====================================================
# MAIN ENTRYPOINT
# =====================================================
def create_excel_dashboard(
    df: pd.DataFrame,
    dashboard_spec: Dict[str, Any],
    template_path: Optional[str],
    output_path: str,
    visible: bool = False,
) -> Path:

    if df is None or df.empty:
        raise ValueError("Empty DataFrame")

    template = Path(template_path)
    cfg = read_config(template)

    # Copy template first
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    shutil.copy2(template, tmp)

    # Inject Master sheet
    with pd.ExcelWriter(tmp, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        df.to_excel(w, sheet_name=MASTER_SHEET, index=False)

    excel = open_excel(visible)
    try:
        wb = excel.Workbooks.Open(str(tmp.resolve()))
        dash = wb.Worksheets(DASHBOARD_SHEET)
        master = wb.Worksheets(MASTER_SHEET)

        try:
            last = master.Cells.SpecialCells(XL_CELL_TYPE_LAST_CELL)
            last_row, last_col = last.Row, last.Column
        except Exception:
            last_row = master.UsedRange.Rows.Count
            last_col = master.UsedRange.Columns.Count

        src = f"'{MASTER_SHEET}'!$A$1:${get_column_letter(last_col)}${last_row}"
        cache = create_pivot_cache(wb, src)

        pivots = {
            p["name"]: add_pivot(wb, cache, p)
            for p in dashboard_spec["pivots"]
        }

        placeholders: Dict[str, Tuple[int, int]] = {}
        for r in range(1, 100):
            for c in range(1, 50):
                v = dash.Cells(r, c).Value
                if isinstance(v, str) and v.upper().startswith("CHART"):
                    placeholders[v.upper()] = (r, c)

        used = set()

        for c in dashboard_spec["charts"]:
            ph = c["placeholder"].upper()
            r, col = placeholders[ph]
            used.add(ph)

            chart = dash.ChartObjects().Add(
                dash.Cells(r, col).Left,
                dash.Cells(r, col).Top,
                CHART_WIDTH,
                CHART_HEIGHT,
            ).Chart

            key = str(c["type"]).lower().replace(" ", "").replace("_", "")
            chart.ChartType = XL_CHART_TYPES[key]
            chart.SetSourceData(pivots[c["pivot"]].TableRange2)

            style_chart(chart, cfg, c.get("title"))

        for name, (r, c) in placeholders.items():
            if name not in used:
                dash.Cells(r, c).Value = ""

        for name, value in dashboard_spec.get("kpis", {}).items():
            try:
                rng = wb.Names(name).RefersToRange
                rng.Value = value
                style_kpi_cell(rng, value, cfg)
            except Exception:
                pass

        for field in dashboard_spec.get("slicers", []):
            try:
                sc = wb.SlicerCaches.Add(cache, field)
                slicer = sc.Slicers.Add(dash)

                # ---- POSITION & STYLE (CRITICAL) ----
                slicer.Left = dash.Range("L28").Left
                slicer.Top = dash.Range("L28").Top
                slicer.Width = 150
                slicer.Height = 200
                slicer.Style = cfg.get("slicer_style", "SlicerStyleLight1")

            except Exception as e:
                logger.warning("Skipping slicer '%s': %s", field, e)

        out = Path(output_path).resolve()
        out.parent.mkdir(parents=True, exist_ok=True)
        if out.exists():
            out.unlink()

        wb.SaveAs(str(out))
        return out

    finally:
        close_excel(excel)
        if tmp.exists():
            tmp.unlink()
