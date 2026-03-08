"""
ProVA — excel_module/excel_engine.py
Builds a professional Excel dashboard from a DashboardPlan dict.
Uses openpyxl only — no COM, no pywin32, no template file required.
Opens the finished file with os.startfile() after saving.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("ProVA.Excel.Engine")

# ─────────────────────────────────────────────────────────────────
# PALETTE  — professional, not default Excel colours
# ─────────────────────────────────────────────────────────────────
PALETTE = [
    "1B3A6B",   # Deep Navy
    "2A9D8F",   # Teal
    "E76F51",   # Coral
    "F4A261",   # Amber
    "457B9D",   # Steel Blue
    "264653",   # Dark Teal
]

_TITLE_BG   = "1B3A6B"
_KPI_HDR_BG = "1B3A6B"
_KPI_VAL_BG = "F0F4FA"
_KPI_BOX_BG = "FFFFFF"
_ALT_BG     = "F7F9FC"
_BORDER_C   = "C5D1E0"
_WHITE      = "FFFFFF"
_DARK       = "1B2B45"
_SUBTEXT    = "6B7A99"

# Chart sizes in cm
_CW = 15.0
_CH = 10.0

# Dashboard anchor cells for up to 4 charts (row, col-letter)
_CHART_ANCHORS = ["B9", "L9", "B30", "L30"]


# ─────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────

def _fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_col, end_color=hex_col)


def _font(bold: bool = False, size: int = 11,
          color: str = _DARK, name: str = "Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h: str = "center", v: str = "center",
           wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_box() -> Border:
    s = Side(style="thin", color=_BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom_only() -> Border:
    return Border(bottom=Side(style="thin", color=_BORDER_C))


def _set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


# ─────────────────────────────────────────────────────────────────
# CHART BUILDING
# ─────────────────────────────────────────────────────────────────

def _make_chart(ws_cd, chart_spec: Dict, start_row: int,
                n_data_rows: int) -> Any:
    ctype     = chart_spec["type"]
    color     = PALETTE[chart_spec["color_idx"] % len(PALETTE)]
    title     = chart_spec["title"]
    data_col  = 2   # ChartData always: col1=categories, col2=values

    if ctype == "line":
        chart = LineChart()
        chart.style = 10
    else:
        chart = BarChart()
        chart.type     = "bar" if ctype == "bar" else "col"
        chart.grouping = "clustered"
        chart.style    = 10

    chart.title  = title
    chart.width  = _CW
    chart.height = _CH
    chart.legend = None   # single series — legend wastes space

    # Data (include header row for series name)
    data_ref = Reference(
        ws_cd,
        min_col=data_col, max_col=data_col,
        min_row=start_row,
        max_row=start_row + n_data_rows,
    )
    chart.add_data(data_ref, titles_from_data=True)

    # Categories (skip header)
    cat_ref = Reference(
        ws_cd,
        min_col=1,
        min_row=start_row + 1,
        max_row=start_row + n_data_rows,
    )
    chart.set_categories(cat_ref)

    # Remove gridlines — cleaner look
    try:
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
    except Exception:
        pass

    # Apply palette colour to series
    if chart.series:
        s = chart.series[0]
        try:
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.solidFill = color
        except Exception:
            pass
        # For line charts also set line width
        if ctype == "line":
            try:
                s.graphicalProperties.line.width = 20000  # ~1.5pt in EMU
            except Exception:
                pass

    return chart


# ─────────────────────────────────────────────────────────────────
# CHART DATA SHEET
# ─────────────────────────────────────────────────────────────────

def _write_chart_data(ws_cd, charts: List[Dict]) -> List[Tuple[int, int]]:
    """
    Write each chart's aggregated df to ChartData sheet.
    Returns list of (header_row, n_data_rows) per chart.
    """
    positions: List[Tuple[int, int]] = []
    current_row = 1

    for spec in charts:
        cdf     = spec["df"]
        cat_col = spec["cat_col"]
        val_col = spec["val_col"]
        n       = len(cdf)

        # Header
        ws_cd.cell(row=current_row, column=1, value=cat_col)
        ws_cd.cell(row=current_row, column=2, value=val_col)

        # Data rows
        for i, (cat_v, val_v) in enumerate(
            zip(cdf[cat_col].tolist(), cdf[val_col].tolist()), 1
        ):
            ws_cd.cell(row=current_row + i, column=1, value=cat_v)
            val_cell = ws_cd.cell(row=current_row + i, column=2)
            # Write as float if numeric to avoid type issues in charts
            try:
                val_cell.value = float(val_v)
            except (TypeError, ValueError):
                val_cell.value = val_v

        positions.append((current_row, n))
        current_row += n + 3   # gap between tables

    return positions


# ─────────────────────────────────────────────────────────────────
# DATA SHEET
# ─────────────────────────────────────────────────────────────────

def _write_data_sheet(ws_data, df: pd.DataFrame) -> None:
    ws_data.freeze_panes = "A2"

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(row=1, column=ci, value=col_name)
        cell.fill      = _fill(_TITLE_BG)
        cell.font      = _font(bold=True, size=10, color=_WHITE)
        cell.alignment = _align()
        ws_data.column_dimensions[get_column_letter(ci)].width = max(
            14, min(30, len(str(col_name)) + 4)
        )

    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        bg = _ALT_BG if ri % 2 == 0 else _WHITE
        for ci, val in enumerate(row_vals, 1):
            cell = ws_data.cell(row=ri, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.font = _font(size=10)

    _set_row_height(ws_data, 1, 20)


# ─────────────────────────────────────────────────────────────────
# KPI BLOCK
# ─────────────────────────────────────────────────────────────────
# Each KPI is a 3-col-wide merged block: header row + value row
# KPI formulas reference the Data sheet live — stays dynamic.

_KPI_DEFS = [
    ("Total",   lambda col, n: f"=SUM(Data!{col}2:{col}{n+1})"),
    ("Average", lambda col, n: f"=AVERAGE(Data!{col}2:{col}{n+1})"),
    ("Maximum", lambda col, n: f"=MAX(Data!{col}2:{col}{n+1})"),
    ("Minimum", lambda col, n: f"=MIN(Data!{col}2:{col}{n+1})"),
    ("Records", lambda col, n: f"=COUNTA(Data!A2:A{n+1})"),
]

_KPI_NUMBER_FMT = "#,##0.0"
_KPI_INT_FMT    = "#,##0"


def _write_kpi_row(ws_dash, plan: Dict) -> None:
    """Write 5 KPI boxes starting at row 4, using Excel formulas."""
    df_cols    = plan["df_columns"]
    primary    = plan["primary_metric"]
    n_rows     = plan["n_rows"]

    # Find primary metric column letter in Data sheet (1-indexed same as df)
    try:
        col_idx    = df_cols.index(primary) + 1
        col_letter = get_column_letter(col_idx)
    except ValueError:
        col_letter = "B"   # fallback

    # KPI boxes: each spans 3 columns starting at B (cols 2–4, 5–7, 8–10, 11–13, 14–16)
    kpi_start_col = 2
    kpi_gap       = 3   # columns per KPI box

    for idx, (label, formula_fn) in enumerate(_KPI_DEFS):
        start_col = kpi_start_col + idx * (kpi_gap + 1)  # +1 for gap
        end_col   = start_col + kpi_gap - 1
        start_col_l = get_column_letter(start_col)
        end_col_l   = get_column_letter(end_col)

        hdr_row = 4
        val_row = 5

        # Merge header cells
        ws_dash.merge_cells(
            f"{start_col_l}{hdr_row}:{end_col_l}{hdr_row}"
        )
        hdr_cell = ws_dash[f"{start_col_l}{hdr_row}"]
        hdr_cell.value     = label.upper()
        hdr_cell.fill      = _fill(_KPI_HDR_BG)
        hdr_cell.font      = _font(bold=True, size=9, color=_WHITE)
        hdr_cell.alignment = _align()
        _set_row_height(ws_dash, hdr_row, 18)

        # Merge value cells
        ws_dash.merge_cells(
            f"{start_col_l}{val_row}:{end_col_l}{val_row}"
        )
        val_cell = ws_dash[f"{start_col_l}{val_row}"]
        formula  = formula_fn(col_letter, n_rows)
        val_cell.value     = formula
        val_cell.fill      = _fill(_KPI_VAL_BG)
        val_cell.font      = _font(bold=True, size=16, color=_TITLE_BG)
        val_cell.alignment = _align()
        val_cell.number_format = _KPI_INT_FMT if label == "Records" else _KPI_NUMBER_FMT
        _set_row_height(ws_dash, val_row, 36)

        # Box border around both rows
        for row in (hdr_row, val_row):
            for col in range(start_col, end_col + 1):
                ws_dash.cell(row=row, column=col).border = _border_box()

    # Label: which metric are KPIs about
    label_col = get_column_letter(kpi_start_col)
    lbl_cell  = ws_dash[f"{label_col}7"]
    lbl_cell.value     = f"↑  Metric: {primary}"
    lbl_cell.font      = _font(size=9, color=_SUBTEXT)
    lbl_cell.alignment = _align(h="left")


# ─────────────────────────────────────────────────────────────────
# TITLE BAR
# ─────────────────────────────────────────────────────────────────

def _write_title(ws_dash, plan: Dict) -> None:
    source  = plan["source_name"]
    today   = datetime.now().strftime("%d %b %Y")
    title   = f"  {source}  —  Analytics Dashboard"
    subtitle = f"Generated {today} by ProVA"

    # Title row 1 — full width merge B1:T1
    ws_dash.merge_cells("B1:T1")
    cell = ws_dash["B1"]
    cell.value     = title
    cell.fill      = _fill(_TITLE_BG)
    cell.font      = _font(bold=True, size=16, color=_WHITE)
    cell.alignment = _align(h="left", v="center")
    _set_row_height(ws_dash, 1, 36)

    # Subtitle row 2
    ws_dash.merge_cells("B2:T2")
    cell2 = ws_dash["B2"]
    cell2.value     = subtitle
    cell2.fill      = _fill("243F6E")  # slightly lighter navy
    cell2.font      = _font(size=9, color="A8C4E8")
    cell2.alignment = _align(h="left", v="center")
    _set_row_height(ws_dash, 2, 18)

    # Spacer row 3
    ws_dash.merge_cells("B3:T3")
    ws_dash["B3"].fill = _fill(_WHITE)
    _set_row_height(ws_dash, 3, 8)


# ─────────────────────────────────────────────────────────────────
# COLUMN WIDTHS
# ─────────────────────────────────────────────────────────────────

def _set_column_widths(ws_dash) -> None:
    ws_dash.column_dimensions["A"].width = 2   # narrow left margin
    for col_idx in range(2, 26):
        ws_dash.column_dimensions[get_column_letter(col_idx)].width = 9


# ─────────────────────────────────────────────────────────────────
# MAIN BUILD FUNCTION
# ─────────────────────────────────────────────────────────────────

def build_dashboard(
    df: pd.DataFrame,
    plan: Dict[str, Any],
    output_path: str,
) -> Path:
    """
    Build a complete Excel dashboard from df + plan.
    Saves to output_path and opens it in Excel.
    Returns the output Path.
    """
    if df is None or df.empty:
        raise ValueError("Cannot build dashboard: DataFrame is empty.")

    wb = Workbook()

    # ── Sheets ──────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False   # cleaner dashboard look

    ws_cd = wb.create_sheet("ChartData")
    ws_cd.sheet_state = "hidden"

    ws_data = wb.create_sheet("Data")

    # ── Write raw data ───────────────────────────────────────────
    _write_data_sheet(ws_data, df)

    # ── Write chart data + record positions ─────────────────────
    positions = _write_chart_data(ws_cd, plan["charts"])

    # ── Build charts ─────────────────────────────────────────────
    charts_built = []
    for i, (spec, (start_row, n_data_rows)) in enumerate(
        zip(plan["charts"], positions)
    ):
        try:
            chart = _make_chart(ws_cd, spec, start_row, n_data_rows)
            charts_built.append((chart, _CHART_ANCHORS[i]))
        except Exception as e:
            log.warning("Chart %d (%s) failed: %s", i + 1, spec.get("title"), e)

    # ── Build Dashboard sheet ────────────────────────────────────
    _set_column_widths(ws_dash)
    _write_title(ws_dash, plan)
    _write_kpi_row(ws_dash, plan)

    # Section header for charts
    ws_dash.merge_cells("B8:T8")
    sec = ws_dash["B8"]
    sec.value     = "  Charts"
    sec.fill      = _fill("EEF2F8")
    sec.font      = _font(bold=True, size=10, color=_TITLE_BG)
    sec.alignment = _align(h="left")
    sec.border    = _border_bottom_only()
    _set_row_height(ws_dash, 8, 18)

    for chart, anchor in charts_built:
        ws_dash.add_chart(chart, anchor)

    # Row heights for chart areas (prevent compressed look)
    for row in range(9, 55):
        _set_row_height(ws_dash, row, 15)

    # ── Save ─────────────────────────────────────────────────────
    out = Path(output_path).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists():
        try:
            out.unlink()
        except Exception:
            pass
    wb.save(str(out))
    log.info("Dashboard saved → %s", out)

    # ── Open in Excel ────────────────────────────────────────────
    try:
        os.startfile(str(out))
    except Exception as e:
        log.warning("Could not auto-open dashboard: %s", e)

    return out